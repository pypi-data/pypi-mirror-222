# ChatGPT <3
import os

import openpyxl
import pytest
from koalak.utils import data_to_excel
from openpyxl import Workbook


@pytest.fixture
def data():
    return [
        {"name": "John Doe", "age": 30, "email": "johndoe@example.com"},
        {"name": "Jane Smith", "age": 25, "email": "janesmith@example.com"},
        {"name": "Bob Johnson", "age": 40, "email": "bobjohnson@example.com"},
    ]


def test_data_to_excel_returns_workbook(data):
    wb = data_to_excel(data)
    assert isinstance(wb, Workbook)


def test_new_workbook(data):
    # Test creating a new workbook with default sheet
    wb = data_to_excel(data)
    assert len(wb.sheetnames) == 1
    assert wb.sheetnames[0] == "data"


def test_new_workbook_sheet_name(data):
    # Test creating a new workbook with a specified sheet name
    wb = data_to_excel(data, sheet_name="test")
    assert len(wb.sheetnames) == 1
    assert wb.sheetnames[0] == "test"


def test_existing_workbook_sheet_name(data):
    # Test adding a sheet to an existing workbook with a specified sheet name
    wb = Workbook()
    ws = wb.active
    ws.title = "old_sheet"
    data_to_excel(data, sheet_name="new_sheet", workbook=wb)
    assert len(wb.sheetnames) == 2
    assert wb.sheetnames[0] == "old_sheet"
    assert wb.sheetnames[1] == "new_sheet"


def test_save_file(data, tmp_path):
    # Test saving the workbook to a file
    filename = tmp_path / "test.xlsx"
    data_to_excel(data, save=str(filename))
    wb = openpyxl.load_workbook(str(filename))
    assert len(wb.sheetnames) == 1
    assert wb.sheetnames[0] == "data"


def test_save_file_sheet_name(data, tmp_path):
    # Test saving the workbook to a file with a specified sheet name
    filename = tmp_path / "test.xlsx"
    data_to_excel(data, sheet_name="test", save=str(filename))
    wb = openpyxl.load_workbook(str(filename))
    assert len(wb.sheetnames) == 1
    assert wb.sheetnames[0] == "test"


def test_data_to_excel_basic(tmp_path, data):
    # Create a file path for the output Excel file
    excel_file = str(tmp_path / "test_data.xlsx")

    # Create an Excel table from the test data
    workbook1 = data_to_excel(data, save=excel_file)
    workbook2 = openpyxl.load_workbook(excel_file)

    for workbook in [workbook1, workbook2]:
        # Assert that the workbook was returned
        assert isinstance(workbook, Workbook)

        # Assert that the workbook contains one sheet
        assert len(workbook.sheetnames) == 1

        # Get a reference to the worksheet
        ws = workbook.active

        # Assert that the worksheet name matches the default
        assert ws.title == "data"

        # Assert that the headers were written correctly
        assert [cell.value for cell in ws[1]] == ["name", "age", "email"]

        # Assert that the first row of data was written correctly
        assert [cell.value for cell in ws[2]] == ["John Doe", 30, "johndoe@example.com"]

        # Assert that the second row of data was written correctly
        assert [cell.value for cell in ws[3]] == [
            "Jane Smith",
            25,
            "janesmith@example.com",
        ]

        # Assert that the third row of data was written correctly
        assert [cell.value for cell in ws[4]] == [
            "Bob Johnson",
            40,
            "bobjohnson@example.com",
        ]

        # Assert that the table display name matches the worksheet name
        assert "data" in ws.tables

        # Assert that the table range matches the worksheet range
        assert ws.tables["data"].ref == "A1:C4"

        # Assert that the file was saved to disk
        assert os.path.exists(excel_file)


def test_data_to_excel_existing_workbook(data):
    # Create an Excel table from the test data
    workbook1 = data_to_excel(data, sheet_name="data1")
    workbook2 = data_to_excel(data, workbook=workbook1, sheet_name="data2")

    for sheetname in ["data1", "data2"]:
        ws = workbook2[sheetname]

        # Assert that the worksheet name matches the default
        assert ws.title == sheetname

        # Assert that the headers were written correctly
        assert [cell.value for cell in ws[1]] == ["name", "age", "email"]

        # Assert that the first row of data was written correctly
        assert [cell.value for cell in ws[2]] == ["John Doe", 30, "johndoe@example.com"]

        # Assert that the second row of data was written correctly
        assert [cell.value for cell in ws[3]] == [
            "Jane Smith",
            25,
            "janesmith@example.com",
        ]

        # Assert that the third row of data was written correctly
        assert [cell.value for cell in ws[4]] == [
            "Bob Johnson",
            40,
            "bobjohnson@example.com",
        ]
