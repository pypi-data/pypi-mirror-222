import csv
import io
import itertools
import os
import random
import shutil
import string
import sys
import tempfile
import types
from contextlib import contextmanager
from typing import Any, Dict, Iterable, List, Optional, Tuple, Union

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table


@contextmanager
def tmp_module(
    module_name: str = None,
    context: Dict[str, Any] = None,
    addsys: bool = True,
    overwrite=False,
):
    """Generate a temporal module as context.
    Temporally add the module to sys.modules so it can be importable
    Args:
    """
    # TODO: unit test this function
    if not module_name:
        module_name = "tmpmodule_" + randomstr(exclude=list(sys.modules.keys()))
    else:
        if module_name in sys.modules and not overwrite:
            raise TabError("Module name already exist, call it with overwrite = True")

    context = context or {}
    module = types.ModuleType(module_name)
    module.__dict__.update(context)
    sys.modules[module_name] = module
    yield module
    del sys.modules[module_name]


@contextmanager
def temp_pathname():
    """Return an available pathname that you can use for files/dirs

    Will automatically clean the pathname if it's
    a file or directory"""
    tmpdir = tempfile.gettempdir()
    pathname = os.path.join(
        tmpdir, randomstr(20, prefix="tmppathname_", exclude=os.listdir(tmpdir))
    )
    try:
        yield pathname
    finally:
        if os.path.isdir(pathname):
            shutil.rmtree(pathname)
        elif os.path.isfile(pathname):
            os.remove(pathname)


def randomstr(
    n: int = 10,
    *,
    prefix="",
    suffix="",
    alphabet: str = string.ascii_letters,
    exclude: Iterable = None,
):
    """Generate random string of length n
    Args:
        n: length of returned string (without the prefix)
        alphabet: alphabet to use
        exclude: generated string must not be in this list (or any iterable object)
    """
    if exclude is not None:
        exclude = set(exclude)
    else:
        exclude = set()

    while True:
        string = prefix + "".join(random.choices(alphabet, k=n)) + suffix
        if string not in exclude:
            return string


@contextmanager
def temp_str2filename(string: str) -> str:
    """Generate a temporary file with the string content and
    return it's filename
    """
    with temp_pathname() as pathname:
        with open(pathname, "w") as f:
            f.write(string)
        yield pathname


@contextmanager
def str2stdin(string: str):
    string_io = io.StringIO(string)
    old_stdin = sys.stdin
    sys.stdin = string_io
    yield
    sys.stdin = old_stdin


@contextmanager
def file2stdin(filename: str):
    with open(filename) as f:
        old_stdin = sys.stdin
        sys.stdin = f
        try:
            yield
        # FIXME unit test finally with all other contextmanager
        finally:
            sys.stdin = old_stdin


def get_prefixed_callables_of_object(obj, prefix: str):
    """Get all the methods with a specific prefix of an object
    Example:
        get_methods_with_prefix(obj, "test_")
    """
    return [
        getattr(obj, e)
        for e in dir(obj)
        if e.startswith(prefix) and callable(getattr(obj, e))
    ]


def humaintime(time: float) -> str:
    """Return the given bytes as a human friendly KB, MB, GB, or TB string."""
    if time < 60:
        label = "seconds"
    elif time < 3600:  # 60 ** 2
        label = "minutes"
        time /= 60
    elif time < 216000:  # 60 ** 3
        label = "hours"
        time /= 3600
    else:
        label = "days"
        time /= 86400
    return f"{time:.2f} {label}"


humantime = humaintime


def humanbytes(size: int) -> str:
    """Return the given bytes as a human friendly KB, MB, GB, or TB string."""
    if size < 1024:
        label = "Bytes"
    elif size < 1_048_576:  # 1024 ** 2
        label = "KB"
        size /= 1024
    elif size < 1_073_741_824:  # 1024 ** 3
        label = "MB"
        size /= 1_048_576
    else:
        label = "GB"
        size /= 1_073_741_824

    if label == "Bytes":
        return f"{size} Bytes"
    else:
        return f"{size:.2f} {label}"


def normalize_optional_list_of_str(obj):
    if isinstance(obj, str):
        return [obj]
    elif obj is None:
        return []
    return obj


def str_find_all(
    text: str,
    substring: str,
    start: int = None,
    end: int = None,
    overlap: bool = None,
    return_position: bool = None,
) -> Union[List[int], List[Tuple[int, int]]]:
    """
    Finds all indexes or positions of a substring in a given text.

    Args:
        text: The text to search for the substring in.
        substring: The substring to search for in the text.
        start: The starting index of the search. Defaults to None.
        end: The ending index of the search. Defaults to None.
        overlap: A boolean indicating whether overlapping occurrences of the substring should be included.
            If not specified, defaults to False.
        return_position: A boolean indicating whether to return a list of indexes or a list of tuples representing
            the start and end positions of each occurrence of the substring in the text. If not specified, defaults to False.

    Returns:
        A list of integer indexes or a list of tuples representing the start and end positions of each occurrence of the
        substring in the text. If the substring is not found, an empty list is returned.

    Example:
        >>> str_find_all("hello world", "l")
        [2, 3, 9]
        >>> str_find_all("hello world", "l", return_position=True)
        [(2, 3), (3, 4), (9, 10)]
    """
    if start is None:
        start = 0
    if end is None:
        end = len(text)
    if overlap is None:
        overlap = False
    if return_position is None:
        return_position = False

    results = []
    while True:
        index = text.find(substring, start, end)
        if index == -1:
            break
        if return_position:
            end_index = index + len(substring)
            results.append((index, end_index))
        else:
            results.append(index)
        if overlap:
            start = index + 1
        else:
            start = index + len(substring)
    return results


# TODO
def str_split_and_keep(text: str, split):
    pass


def str_remove_first_char(text: str, n=1):
    return "\n".join(e[n:] for e in text.split("\n"))


def merge_intervals(intervals: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
    """
    Merge overlapping intervals in a list of intervals.

    Args:
        intervals (List[Tuple[int, int]]): A list of tuples representing intervals, where each tuple contains
            the start and end positions of the interval.

    Returns:
        List[Tuple[int, int]]: A new list of tuples representing merged intervals, where each tuple contains
            the start and end positions of the merged interval.

    Examples:
        >>> intervals = [(1, 3), (2, 6), (8, 10), (15, 18)]
        >>> merged_intervals = merge_intervals(intervals)
        >>> print(merged_intervals)
        [(1, 6), (8, 10), (15, 18)]
    """
    if not intervals:
        return []

    intervals = sorted(intervals, key=lambda x: x[0])
    merged_intervals = [intervals[0]]

    for interval in intervals[1:]:
        if interval[0] <= merged_intervals[-1][1]:
            merged_intervals[-1] = (
                merged_intervals[-1][0],
                max(interval[1], merged_intervals[-1][1]),
            )
        else:
            merged_intervals.append(interval)

    return merged_intervals


DEFAULT_NB_ROWS_TO_CHECK_TO_DEDUCE_HEADERS = 200


def data_to_excel(
    data: List[Dict[str, Any]],
    *,
    sheet_name: str = None,
    workbook: Optional[Union[str, Workbook]] = None,
    save: Optional[str] = None,
    nb_rows_check: Optional[int] = None,
) -> Workbook:
    """
    Creates an Excel table from a list of dictionaries.

    Args:
        data: A list of dictionaries, where each dictionary represents a row
            in the table.
        sheet_name: The name of the worksheet to create. If None, defaults to
            'data'.
        workbook: An optional string or openpyxl.Workbook object. If None, a new
            Workbook will be created. If a string, the Workbook will be loaded
            from the file with that name. If an openpyxl.Workbook object, a new
            worksheet will be added to that workbook.
        save: An optional string representing the file path to save the
            Workbook to. If None, the Workbook will not be saved to a file.

    Returns:
        The openpyxl.Workbook object representing the created Excel table.
    """
    # Check if sheet_name is None and set it to 'data' if it is
    if sheet_name is None:
        sheet_name = "data"

    if nb_rows_check is None:
        nb_rows_check = DEFAULT_NB_ROWS_TO_CHECK_TO_DEDUCE_HEADERS
    elif nb_rows_check == 0:
        try:
            nb_rows_check = len(data)
        except TypeError:
            data = list(data)
            nb_rows_check = len(data)

    data = iter(data)
    data_list = list(itertools.islice(data, nb_rows_check))
    data_iter = itertools.chain(data_list, data)

    # If no workbook is provided, create a new one and remove default sheet named 'Sheet'
    if workbook is None:
        wb = Workbook()
        wb.remove(wb["Sheet"])
    elif isinstance(workbook, str):
        # If a string is provided, load the Workbook from the file
        wb = load_workbook(workbook)
    else:
        # Otherwise, use the provided Workbook and create a new worksheet
        wb = workbook
    ws = wb.create_sheet(title=sheet_name)

    # Write the headers to the worksheet
    headers = list(dict.fromkeys(key for d in data_list for key in d.keys()))
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write the data to the worksheet
    for row_num, row_data in enumerate(data_iter, 2):
        for key in row_data.keys():
            # TODO: test when data are not normalized!
            col_num = headers.index(key) + 1
            ws.cell(row=row_num, column=col_num, value=row_data[key])

    # Create a table from the data
    table = Table(displayName=sheet_name, ref=ws.dimensions)
    ws.add_table(table)

    # Save the workbook if a save path is provided
    if save is not None:
        wb.save(save)

    return wb


def data_to_csv(
    data: Iterable[Dict[str, Any]],
    file_path: str,
    *,
    delimiter: str = ",",
    nb_rows_check: Optional[int] = None,
) -> None:
    """
    Creates a CSV file from a list of dictionaries.

    Args:
        data: A list of dictionaries, where each dictionary represents a row
            in the CSV file.
        file_path: The path to the CSV file to create.
        delimiter: The delimiter to use between fields. Defaults to ','.
        nb_rows_check: The number of rows to check to deduce the headers. If None,
            all rows are checked.
    """

    # If nb_rows_check is None or greater than the number of rows, use all rows
    if nb_rows_check is None:
        nb_rows_check = DEFAULT_NB_ROWS_TO_CHECK_TO_DEDUCE_HEADERS
    elif nb_rows_check == 0:
        try:
            nb_rows_check = len(data)
        except TypeError:
            data = list(data)
            nb_rows_check = len(data)

    data = iter(data)
    data_list = list(itertools.islice(data, nb_rows_check))
    data_iter = itertools.chain(data_list, data)

    # Deduce the headers from the first nb_rows_check rows
    headers = list(dict.fromkeys(key for d in data_list for key in d.keys()))

    # Write the data to the CSV file
    with open(file_path, mode="w", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=headers, delimiter=delimiter)
        writer.writeheader()
        writer.writerows(data_iter)
